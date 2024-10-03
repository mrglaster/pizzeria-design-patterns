import os
import openpyxl
from src.modules.domain.report.report.base.abstract_report import ComplexReport
from src.modules.domain.report.report_format.report_format import ReportFormat
from src.modules.exception.bad_argument_exception import BadArgumentException
from src.modules.validation.data_validator import DataValidator
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


class ReportXLSX(ComplexReport):

    def __init__(self, settings_path=""):
        super().__init__(settings_path=settings_path)
        self.format = ReportFormat.FORMAT_XLSX

    def create(self, data: list):
        DataValidator.validate_field_type(data, list)
        if not len(data):
            raise BadArgumentException("Empty data list provided!")
        used_model = data[0]
        class_name = type(used_model).__name__
        self._document = openpyxl.Workbook()
        sheet = self._document.active
        sheet.title = class_name
        fields = list(filter(lambda x: not x.startswith("_") and not callable(getattr(used_model.__class__, x)),
                             dir(used_model.__class__)))
        for col_idx, field in enumerate(fields, 1):
            column_letter = get_column_letter(col_idx)
            cell = sheet[f"{column_letter}1"]
            cell.value = field
            cell.font = Font(bold=True)

        for row_idx, row in enumerate(data, 2):
            for col_idx, field in enumerate(fields, 1):
                value = getattr(row, field)
                sheet.cell(row=row_idx, column=col_idx, value=str(value))

    def save(self, file_name: str = 'report.xlsx') -> bool:
        try:
            DataValidator.validate_field_type(file_name, str)
            DataValidator.validate_str_not_empty(file_name)
            if self._document is None:
                raise ValueError("No data to save. Please create the report first.")
            save_path = file_name
            if os.path.basename(file_name) == file_name:
                save_path = os.path.join(self.settings_manager.settings.reports_path, file_name)
            self._document.save(save_path)
            return True
        except Exception as e:
            self.exception = e
            return False
